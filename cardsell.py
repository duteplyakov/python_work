# import openpyxl
#
# wb = openpyxl.load_workbook('cardsell.xlsx')
#
# sheets = wb.sheetnames
# for sheet in sheets:
#    print(sheet)
#
# sheet = wb.active
#
# print(sheet['A2'].value)
# print(sheet['B1'].value)
# print(sheet['C1'].value)
# print(sheet['D1'].value)
#
# print()
#


from openpyxl import load_workbook

wb = load_workbook('cardsell.xlsx')
sheet_ranges = wb['Лист1']
column_b = sheet_ranges['D']

for i in range(len(column_b)):
    print(column_b[i].value)

# Или просто как итератор перебрать столбец:
for cell in column_b:
    print(cell.value)
